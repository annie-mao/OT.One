from openpyxl import load_workbook
from openpyxl.styles import colors
from baseProtocol import BaseProtocol, InvalidEntry
from collections import OrderedDict
from pprint import PrettyPrinter
import math

class PCR:
    def __init__(self, name, description, notes):
        self.name = name
        self.description = description
        self.notes = notes
        self.copies = 1
        self.groups = []
        self.cyclerPtr = 0
        self.cyclerPrograms = ["PFUNKEL1","PFUNKEL2","PFUNKEL3"]
        self.excel = {
            'sheet_name': 'For Robot (w Mastermixes)',
            'start_row': 20,
            'reagent': {'key': 'D4', 'scan': 'D', 'name': 'A'},
            'cycler': {'key': 'D5', 'scan': 'B'},
            'aliquot': {'key': 'D6', 'scan': 'D'},
            'unshared' : ['Primary oligos (uM)',' secondary oligo  (uM)',\
                        'ssDNA (ng/uL)']
        } 
        self.mix_settings = {
            "repetitions": 5,
            "blowout": True,
            "liquid-tracking": False
        }
        self.cycler_settings = {
            "control": "CALC",
            "heated-lid": False,
            "vessel": "Tubes",
            "volume": 100
        }
        self.oil_transfer_settings = {
            "extra-pull": True,
            "to": {
                "tip-offset": 10,
                "touch-tip": False
            }
        }
        self.after_oil_transfer_settings = {
            "extra-pull": False,
            "to": {
                "tip-offset": -15,
                "touch-tip": False
            }
        }
        self.after_oil_mix_settings = {
            "tip-offset": -15
        }
        self.aliquot_transfer_settings = {
            "from": {
                "tip-offset": -15
            },
            "to": {
                "tip-offset": 0
            }
        }
        self.minVol = 5
        self.volScaleFactor = 0.1
        self.mixVol = 10
        self.oilVol = 60
        self.aliquotRow = "A"
        self.aliquotVol = 5
        self.aliquotTotalVol = 20
        self.aliquotNum = 1
        self.useOil = True
        self.addedOil = False
        self.useDye = True
        self.holdTemp = 4
        self.waterName = 'DI-NF water'
        self.cyclerContainer = "2-8-tube-strip"
        self.iceContainer = "96-PCR-tubes"
        self.sharedContainerOptions = OrderedDict([
            (600, "tube-strip-600ul"),
            (1500, "tube-strip-1.5ml"),
            (2000, "tube-strip-2ml")
        ])
        self.row = "A"
        self.printer = PrettyPrinter(indent = 2)


    def duplicate(self,copies):
        self.copies = copies


    def set_shared_reagents(self,reagents):
        """input list of reagents to set as shared
        """
        reagents = BaseProtocol.listify(reagents)
        for reagent in reagents:
            self.excel['shared'].append(reagent)

#########################################################################
#-------------------------------- EXCEL ---------------------------------

    def import_from_excel(self,filename):
        wb = load_workbook(filename, data_only = True)
        ws = wb[self.excel['sheet_name']]
        self.excel['reagent']['color'] = \
            ws[self.excel['reagent']['key']].fill.start_color.index
        self.excel['cycler']['color'] = \
            ws[self.excel['cycler']['key']].fill.start_color.index
        self.excel['aliquot']['color'] = \
            ws[self.excel['aliquot']['key']].fill.start_color.index
        row = self.excel['start_row']
        
        # scan down sheet and look for color-coded reagent, cycler,
        # and aliquot groups
        while (row < ws.max_row):
            rCell = self.excel['reagent']['scan'] + str(row)
            cCell = self.excel['cycler']['scan'] + str(row)
            aCell = self.excel['aliquot']['scan'] + str(row)
            rColor = ws[rCell].fill.start_color.index
            cColor = ws[cCell].fill.start_color.index
            aColor = ws[aCell].fill.start_color.index
            if (rColor == self.excel['reagent']['color']):
                row = self.get_reagent_group_from_excel(ws,row)
            elif (cColor == self.excel['cycler']['color']):
                row = self.get_cycler_group_from_excel(ws,row)
            elif (aColor == self.excel['aliquot']['color']):
                row = self.get_aliquot_group_from_excel(ws,row)
            else:
                row = row + 1

    
    def get_reagent_group_from_excel(self,ws,row):
        """ get a group of reagents from excel sheet
        returns the next row after the group
        """
        scanCell = self.excel['reagent']['scan'] + str(row)
        nameCell = self.excel['reagent']['name'] + str(row)
        while ws[scanCell].fill.start_color.index ==\
                            self.excel['reagent']['color']:
            # get reagent name from nameCell
            name = ws[nameCell].value
            reagentGroup = {
                "type": "reagent",
                "name": name,
                # get reagent volume from scanCell
                "volume": ws[scanCell].value,
                # look for reagent in the set of shared reagents
                "shared": (name not in self.excel["unshared"])
            }
            # add reagent group to self.groups
            self.groups.append(reagentGroup)
            # increment to next row
            row = row + 1
            scanCell = self.excel['reagent']['scan'] + str(row)
            nameCell = self.excel['reagent']['name'] + str(row) 
        return row 


    def get_cycler_group_from_excel(self,ws,row):
        """ get a cycler instruction group from excel sheet
        right now just picks from a list of predefined cycler
        program names. Can be extended to create custom cycler
        programs based on the temperature steps and durations.
        """
        scanCell = self.excel['cycler']['scan'] + str(row)
        loadGroup = True
        while ws[scanCell].fill.start_color.index ==\
                            self.excel['cycler']['color']:
            if loadGroup:
                name = self.cyclerPrograms[self.cyclerPtr] 
                cyclerGroup = {
                    "type": "cycler",
                    "name": name
                }
                # add cycler group to self.groups
                self.groups.append(cyclerGroup)
                # don't load a new group in the next iteration of while
                loadGroup = False
                # increment cycler pointer
                self.cyclerPtr=(self.cyclerPtr+1)%len(self.cyclerPrograms)
            # increment to next row
            row = row + 1
            scanCell = self.excel['cycler']['scan'] + str(row)
        return row 


    def get_aliquot_group_from_excel(self,ws,row):
        """ get an aliquot instruction group from excel sheets
        """
        scanCell = self.excel['aliquot']['scan'] + str(row)
        while ws[scanCell].fill.start_color.index ==\
                                self.excel['aliquot']['color']:
            aliquotGroup = {
                "type": "aliquot",
                # get aliquot volume from scanCell
                "volume": ws[scanCell].value,
            }
            # add aliquot group to self.groups
            self.groups.append(aliquotGroup)
            # increment to next row
            row = row + 1
            scanCell = self.excel['aliquot']['scan'] + str(row)
        return row


########################################################################
#---------------------- BASEPROTOCOL CONVERSION ------------------------

    def convert_to_protocol(self):
        """ convert to a baseProtocol class
        """
        # instantiate protocol
        self.p = BaseProtocol(self.name,self.description,self.notes)
        # assign cycler and ice containers
        self.p.assign_labware("ice",self.iceContainer)
        
        # create a reaction tube for each copy of the protocol
        rxnTubes = []
        for i in range(0,self.copies):
            rxnTubes.append("rxn_"+str(i+1))
        # preload reagents
        self.preload_reagents(rxnTubes)
        # preload aliquots
        self.preload_aliquots(rxnTubes)
        # preload cycler programs
        self.preload_cycler(rxnTubes)
        # add instructions to protocol
        self.add_instructions_to_protocol(rxnTubes)

    
    def export_to_JSON(self,fname):
        """ export baseProtocol to JSON
        """
        print("Exporting protocol to " + fname)
        print("Total p10 tips: " + str(self.p.numTips_p10))
        print("Total p200 tips: " + str(self.p.numTips_p200))
        return self.p.export_to_JSON(fname)


    def add_instructions_to_protocol(self,rxnTubes):
        # loop through instruction groups and add to protocol
        for group in self.groups:
            if group["type"] == "reagent" and group["shared"]:
                self.convert_shared_reagent_group(group,rxnTubes)
            elif group["type"] == "reagent" and not group["shared"]:
                self.convert_reagent_group(group,rxnTubes)
            elif group["type"] == "aliquot":
                self.convert_aliquot_group(group,rxnTubes)
            elif group["type"] == "cycler":
                self.convert_cycler_group(group,rxnTubes)
        self.cleanup(rxnTubes)


    def cleanup(self,rxnTubes):
        # assign rxnTubes to cycler container
        for i in range(0,len(rxnTubes)):
            location = self.next_empty_location("cycler")
            self.p.assign_container(rxnTubes[i],"cycler",location)
        # assign shared reagents to shared container
        maxVol = self.findMaxVol(self.p.list_unassigned_locations())
        self.sharedContainer = self.findContainer(maxVol)
        self.p.assign_labware("shared",self.sharedContainer)
        for reagent in self.p.list_unassigned_locations():
            location = self.next_empty_location("shared")
            self.p.assign_container(reagent,"shared",location)


    def findContainer(self,volume):
        """ find the minimum capacity container for storing the given volume
        """
        for containerVol, container in self.sharedContainerOptions.items():
            if containerVol > volume:
                return container


    def findMaxVol(self,locations):
        """ find the maximum volume among a list of locations (e.g. the
        shared reagents)
        """
        maxVol = 0
        for loc in locations:
            if self.p.locations[loc]["maxVol"] > maxVol:
                maxVol = self.p.locations[loc]["maxVol"]
        return maxVol

#---------------------------- REAGENTS ----------------------------------
    def preload_reagents(self,rxnTubes):
        """ Pre-load all reagents into protocol for accurate volume
        calculations
        """
        for group in self.groups:
            if group["type"] == "reagent" and group["shared"]:
                self.add_shared_reagent(group,rxnTubes)
            elif group["type"] == "reagent" and not group["shared"]:
                self.add_reagent(group,rxnTubes)
    

    def add_shared_reagent(self,group,rxnTubes):
        """ add a shared reagent to baseProtocol ingredients dictionary
        """
        name = group["name"]
        volume = group["volume"]
        for i in range(0,len(rxnTubes)):  
            # add ingredient to protocol
            self.p.add_ingredient(name,name,volume)
        self.scale_volumes(name,name,self.p.locations[name]["volume"])


    def add_reagent(self,group,rxnTubes):
        """ add a reagent to baseProtocol ingredients dictionary
        """
        volume = group["volume"]
        container = "ice"
        for i in range(0,len(rxnTubes)): 
            name = group["name"] + str(i+1)
            row = chr(ord(self.row))
            col = str(i+1)
            # add ingredient to protocol 
            self.p.add_ingredient(name,name,group["volume"])
            # assign container
            self.p.assign_container(name,container,row+col)
            self.scale_volumes(name,name,self.p.locations[name]["volume"])
        # increment column for next round of rxn-specific reagents
        self.row = chr(ord(self.row) + 1)


    def convert_shared_reagent_group(self,group,rxnTubes):
        """ convert a reagent group to baseProtocol transfer instructions
        when the reagent is being shared among all rxns
        """
        name = group["name"]
        volume = group["volume"]
        for i in range(0,len(rxnTubes)):  
            # transfer ingredient to rxn tube
            self.p.add_transfer_group(name,rxnTubes[i],volume)


    def convert_reagent_group(self,group,rxnTubes):
        """ convert a reagent group to baseProtocol transfer instructions
        """
        volume = group["volume"]
        for i in range(0,len(rxnTubes)): 
            name = group["name"] + str(i+1)
            # transfer ingredient to rxn tube
            self.p.add_transfer_group(name,rxnTubes[i],group["volume"])


#------------------------------- ALIQUOTS -------------------------------
    def preload_aliquots(self,rxnTubes):
        """ pre-load aliquot reagents to protocol
        """
        for group in self.groups:
            if group["type"] == "aliquot":
                if self.useDye:
                    self.add_dye(rxnTubes,group["volume"])
       
    
    def add_dye(self,rxnTubes,aliquotVol):
        """ load diluted dye for running aliquot gels
        """
        # calculate dye dilution
        dyeVol = self.aliquotTotalVol/6
        waterVol = self.aliquotTotalVol - dyeVol - aliquotVol
        for rxn in rxnTubes:
            # add dye and water to ingredients
            self.p.add_ingredient("dye","dye",dyeVol)
            self.p.add_ingredient(self.waterName,self.waterName,waterVol)
        self.scale_volumes("dye","dye",self.p.locations["dye"]["volume"])
        self.scale_volumes(self.waterName,self.waterName,self.p.locations[self.waterName]["volume"])


    def convert_aliquot_group(self,group,rxnTubes):
        """ convert an aliquot group to baseProtocol instructions
        """
        for i in range(0,len(rxnTubes)):
            name = "aliquot_{0}_{1}".format(str(i+1),self.aliquotNum)
            volume = group["volume"]
            self.p.add_transfer_group(rxnTubes[i],name,volume,self.aliquot_transfer_settings)
            # load dye for gel electrophoresis
            if self.useDye:
                self.transfer_dye(name,volume)
            # assign aliquot to ice container
            row = chr(ord(self.aliquotRow) + len(self.excel["unshared"]))
            col = str(i+1)
            self.p.assign_container(name,"ice",row+col)
        # increment column and # of aliquots for next round
        self.aliquotRow = chr(ord(self.aliquotRow) + 1)
        self.aliquotNum = self.aliquotNum + 1


    def transfer_dye(self, aliquotName, aliquotVol):
        """ load diluted dye for running aliquot gels
        """
        # calculate dye dilution
        dyeVol = self.aliquotTotalVol/6
        waterVol = self.aliquotTotalVol - dyeVol - aliquotVol
        # transfer dye and water to aliquot, mix at end
        self.p.add_transfer_group("dye",aliquotName,dyeVol)
        self.p.add_transfer_with_mix(self.waterName,aliquotName,waterVol)


#------------------------------- CYCLER ---------------------------------
    def preload_cycler(self,rxnTubes):
        """ preload cycler programs in the protocol and make oil
        available as a reagent if using PCR oil
        """
        # assign the "cycler" container (i.e. where all the rxn tubes
        # will be located) to a labware
        self.p.assign_labware("cycler",self.cyclerContainer)
        # add cycler programs to protocol library
        for prog in self.cyclerPrograms:
            self.p.add_cycler_prog(prog,self.holdTemp)
        # if using PCR oil, add to protocol ingredients section
        if self.useOil:
            for rxn in rxnTubes:
                self.p.add_ingredient("oil","oil",self.oilVol)
        self.scale_volumes("oil","oil",self.p.locations["oil"]["volume"])


    def convert_cycler_group(self,group,rxnTubes):
        """ convert a cycler group to baseProtocol instructions
        """
        # mix rxn tubes
        for rxn in rxnTubes: 
            self.p.add_mix_group(rxn,self.mixVol)
        # add PCR oil if necessary (only add oil once, before 1st
        # cycler group)
        if self.useOil and not self.addedOil:
            self.transfer_oil(rxnTubes)
        # add cycler instruction to protocol
        self.p.add_cycler_group(group["name"])


    def transfer_oil(self,rxnTubes):
        """ add oil to PCR tubes before running a cycler program
        to prevent evaporation
        """
        # change default transfer settings for oil settings
        for key, value in self.oil_transfer_settings.items():
            self.p.configure_transfer(key,value)
        # add oil to the top of each rxn tube
        for rxn in rxnTubes:
            self.p.add_transfer_group("oil",rxn,self.oilVol)
        # change default settings for subsequent instructions
        for key,value in self.after_oil_transfer_settings.items():
            self.p.configure_transfer(key,value)
        for key,value in self.after_oil_mix_settings.items():
            self.p.configure_mix(key,value)
        # only add oil once to rxn tubes
        self.addedOil = True


#---------------------------- OTHER HELPERS -----------------------------
    def scale_volumes(self,reagent,location,volume):
        """scale starting volumes of reagents up to account for
        dead volume, reduce pipetting errors. If volume is below
        minimum volume threshold, set to self.minVol
        """
        if volume < self.minVol:
            addVol = self.minVol - volume
        else:
            # round up final volume to nearest whole number
            addVol = self.volScaleFactor*volume
            addVol = addVol + (math.ceil(volume+addVol)-(volume+addVol))
        self.p.add_ingredient(reagent,location,addVol)


    def next_empty_location(self,containerName):
        """ return the next empty location in a container
        e.g. A5
        """
        return self.p.deck[containerName]["empty"][0]

    def next_location(self):
        """ next location in the 96-well PCR plate, according to how
        many rxn-specific locations there are  (min 4 wells needed
        for primary oligo, secondary oligo, ssDNA template, and 1
        aliquot) 
        if 4 wells, allow wraparound to other half of plate
        """
        pass 
