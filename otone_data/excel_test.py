from openpyxl import load_workbook
from openpyxl.styles import colors

wb = load_workbook('robot_protocol_09_29.xlsx', data_only = True)
sheet = wb['Combinatorial (no delay)']
ssDNA = sheet['B3'].value
oligo1 = sheet['B10'].value
oligo2 = sheet['B14'].value

reagentColor = sheet['D4'].fill.start_color.index
cyclerColor = sheet['D5'].fill.start_color.index
aliquotColor = sheet['D6'].fill.start_color.index

row = 20
reagents = True
reagentNames = []
reagentVols = []
looking = True
while reagents:
    if (sheet['D'+str(row)].fill.start_color.index is reagentColor):
        reagentNames.append(sheet['A'+str(row)].value)
        reagentVols.append(float(sheet['D'+str(row)].value))
        row = row + 1
    else:
        reagents = False
        print(reagentNames)
        print(reagentVols)
while looking:
    if (sheet['B'+str(row)].fill.start_color.index is cyclerColor):
        print('found cycler ' + str(row))
        print(sheet['B'+str(row)].value)
        looking = False
    row = row + 1
print('max rows: ' + str(sheet.max_row))    
print('max cols: ' + str(sheet.max_column))
print('B27: ' + str(sheet['B27'].fill.start_color.index))
print('B28: ' + str(sheet['B28'].fill.start_color.index))
print('B29: ' + str(sheet['B29'].fill.start_color.index))
print('B30: ' + str(sheet['B30'].fill.start_color.index))
print('B35: ' + str(sheet['B35'].fill.start_color.index))
